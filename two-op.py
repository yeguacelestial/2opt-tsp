# 2-OP Heuristic
# TODO: Export all steps by step of a)
# a) Best Found Strategy
#    1. Generate Neighborhood of all possible solutions => N(X)
#    2. Evaluate ALL Objective Functions on each of the neighbors on N(X) => f(Y) for each Y in N(X)
#    3. y* = arg min f(Y)
#    4. If f(Y*) < f(X), MOVE to Y* (Better said: update Tour to Y*), and go to Step 1
#    5. Else If f(Y*) >= f(X), STOP
# b) First Found Strategy

import sys
import openpyxl
from itertools import permutations
from itertools import combinations

def main():
    try:
        tour_file = sys.argv[1]
        tour = read_tour(tour_file)
        matrix_xl_file = sys.argv[2]

        # a) Best Found Strategy
        best_found(tour, matrix_xl_file)

        # b) First Found strategy

    except Exception as e:
        raise


def read_tour(filename):
    try:
        tour = []

        with open(f'{filename}', 'r') as f:
            line = list(f.read())

            for c in line:
                try:
                    tour.append(int(c))
                except:
                    pass

            return tour

    except:
        print("[-] Couldn't read file.")


def best_found(tour, exl):
    # output_str = "\na) BEST FOUND TRATEGY IN 2-OPT\n"
    output_str = f'T = {tour}\n'

    list_of_combinations = get_combinations(tour)
    edges = combinations_to_edges(list_of_combinations)

    # Compute current objective function
    current_obj = current_objective(tour, exl)
    output_str += f'Current f(T) = {current_obj}'
    print (output_str)
    
    # 2-OPT algorithm
    two_opt_bf(tour, exl)

    return


def first_found(tour):
    return


def two_opt_bf(tour, exl):
    list_of_combinations = get_combinations(tour)
    edges = combinations_to_edges(list_of_combinations)
    cur_obj_f = current_objective(tour, exl)

    print(f'\n**AVAILABLE EDGES**')
    obj_f_list = []
    moves_dict = {}
    tours_dict = {}
    for edge in edges:
        a = edge
        print(f'\n*a = {a}')
        print(f'T = {tour}')
        edges_nonadj = []
        
        for e in edges:
            if (e[0] not in a) and (e[1] not in a):
                edges_nonadj.append(e)
            else:
                pass
        print(f'NON-ADJACENT EDGES: {edges_nonadj}')

        for na_edge in edges_nonadj:
            # Move(edge, na_edge) => new Tour
            new_tour = two_opt(tour, edge, na_edge)
            objective_function = current_objective(new_tour, exl)
            print(f'Move({edge},{na_edge}) => {new_tour} => f(T) = {objective_function}')

            moves_dict.update([((edge, na_edge), objective_function)])
            tours_dict.update([(tuple(new_tour), objective_function)])
            obj_f_list.append(objective_function)

    min_obj_f = min(obj_f_list)
    print(f"\n**MIN OBJECTIVE FUNCTION => {min_obj_f}")

    min_obj_f_moves = get_key(min_obj_f, moves_dict)
    print(f'**MOVE OF THE MIN OBJECTIVE FUNCTION => {min_obj_f_moves}')

    # a = min_obj_f_moves[0]
    # b = min_obj_f_moves[1]
    # deltaT = compute_delta(a, b, exl)
    # output_str += f'DeltaT = {deltaT}\n'

    min_obj_f_tour = list(get_key(min_obj_f, tours_dict))
    print(f'**TOUR OF THE MIN OBJECTIVE FUNCTION => {min_obj_f_tour}')

    if min_obj_f < cur_obj_f:
        print(f'{min_obj_f} is better than {cur_obj_f}. Updating T...\n')
        tour = min_obj_f_tour.copy()
        best_found(tour, exl)

    elif min_obj_f >= cur_obj_f:
        print("\n\n**LOCAL BEST SOLUTION**")
        print(f"*f(x) = {cur_obj_f}")
        print(f"**TOUR = {tour}")    
    return


def two_opt(tour:list, edge_a:tuple, edge_b:tuple):
    new_tour = tour.copy()
    for index, c in enumerate(new_tour):
        i = edge_a[0]
        j = edge_a[1]
        k = edge_b[0]
        l = edge_b[1]

        if c == i:
            new_tour[index] = i
        elif c == j:
            new_tour[index] = k
        elif c == k:
            new_tour[index] = j
        elif c == l:
            new_tour[index] = l

    return new_tour


def distance(exl, i, j):
    # Open Workbook object
    wb_obj = openpyxl.load_workbook(exl)

    # Get workbook active sheet object from active attribute
    sheet_obj = wb_obj.active

    # Cell object
    cell_obj = sheet_obj.cell(row = i, column = j)

    if cell_obj.value == None:
        cell_obj = sheet_obj.cell(row = j, column = i)

    distance_value = cell_obj.value

    return distance_value


def obj_f(distances:list):
    objective = 0
    for d in distances:
        objective += d
    
    return objective


def current_objective(tour, exl):
    list_of_combinations = get_combinations(tour)
    edges = combinations_to_edges(list_of_combinations)
    distances = []

    for comb in edges:
        d = distance(exl, comb[0], comb[1])
        distances.append(d)
    
    output_str = f'f(T) = '
    for comb in edges:
        if comb == edges[-1]:
            output_str += f'd({comb[0]},{comb[1]})'
        else:
            output_str += f' d({comb[0]},{comb[1]}) + '

    output_str += f'\nf(T) = '
    for d in distances:
        if d == distances[-1]:
            output_str += f'{d}'
        else:
            output_str += f'{d} + '

    objective_function = obj_f(distances)

    output_str += f'\nf(T) = {objective_function}\n'

    return objective_function


def get_key(value, dict):
    for k, v in dict.items():
        if value == v:
            return k

    return "Key doesn't exist"

def compute_delta(a,b, exl):
    i = a[0]
    j = a[1]
    k = b[0]
    l = b[1]

    delta = (distance(exl, i, k) + distance(exl, j, l)) - (distance(exl, i, j) + distance(exl, k, l))

    return delta


def get_combinations(tour):
    comb = combinations(tour, 2)
    comb_list = list(comb)

    return comb_list


def combinations_to_edges(combinations):
    comb_copy = combinations
    first_index_list = []
    edges = []

    # Add all combinations
    for comb in comb_copy:
        if comb[0] not in first_index_list:
            edges.append(comb)
            first_index_list.append(comb[0])

        else:
            pass
    
    # Add last combination
    last_index_list = []
    for comb in comb_copy:
        last_index_list.append(comb[1])
    
    max_index = max(last_index_list)
    first_index = first_index_list[0]

    last_combination = (max_index, first_index)
    edges.append(last_combination)

    return edges


if __name__ == '__main__':
    main()